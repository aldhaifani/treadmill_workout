from flask import Flask, render_template, session
import openpyxl


app = Flask(__name__)
app.secret_key = "HdraYnfHEiTKsR3kFJbfnMdLAfUYGiew"

wb = openpyxl.load_workbook("./static/data_sheets/treadmill_workout.xlsx")
ws = wb.active

data_obj = [
    {
        "label": ws.cell(row=row, column=1).value,
        "time": ws.cell(row=row, column=2).value,
        "incline": ws.cell(row=row, column=3).value,
        "speed": ws.cell(row=row, column=4).value,
    }
    for row in range(2, ws.max_row + 1)
]


@app.route("/")
def home():
    session.clear()
    return render_template("home.html")


@app.route("/exercise")
def exercise():
    try:
        temp = session["roundCounter"]
    except:
        session["roundCounter"] = 0

    try:
        temp = session["data_n"]
    except:
        session["data_n"] = len(data_obj)

    if session["data_n"] > session["roundCounter"]:
        session["label"] = data_obj[session["roundCounter"]]["label"]
        session["time"] = data_obj[session["roundCounter"]]["time"]
        session["incline"] = data_obj[session["roundCounter"]]["incline"]
        session["speed"] = data_obj[session["roundCounter"]]["speed"]
        session["roundCounter"] = session["roundCounter"] + 1

        return render_template(
            "exercise.html",
            label=session["label"],
            time=session["time"],
            incline=session["incline"],
            speed=session["speed"],
        )
    else:
        render_template("index.html")


if __name__ == "__main__":
    app.run(debug=True)
